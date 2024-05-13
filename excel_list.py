from openpyxl import Workbook
from openpyxl.styles import Alignment
import db_select

import sqlite3
from collections import defaultdict

def mega_fun():
    def select_from_bd(s_d, e_d):
        conn = sqlite3.connect('asd.db')
        cursor = conn.cursor()

        cursor.execute('SELECT tovar, kolvo, price*kolvo FROM nomenkl WHERE rashod == 0 AND data BETWEEN ? AND ?', (s_d, e_d))

        prihod = cursor.fetchall()

        cursor.execute('SELECT tovar, kolvo, price*kolvo FROM nomenkl WHERE rashod == 1 AND data BETWEEN ? AND ?', (s_d, e_d))

        rashod = cursor.fetchall()

        cursor.execute('SELECT tovar, kolvo, price*kolvo FROM nomenkl WHERE rashod == 0 AND data < ?', (s_d,))

        nach = cursor.fetchall()

        conn.close()

        return nach, prihod, rashod

    def sebest():
        conn = sqlite3.connect('asd.db')
        cursor = conn.cursor()

        cursor.execute('SELECT tovar, price FROM nomenkl WHERE rashod == 0')

        asdqwe = cursor.fetchall()

        conn.close()

        return asdqwe

    data = select_from_bd('2006-01-03', '2006-01-05')

    sebestoimost = sebest()

    print(sebestoimost)

    print(data)

    tovari_m = set()

    nach_o = defaultdict(lambda: [0, 0])
    prihod = defaultdict(lambda: [0, 0])
    rashod = defaultdict(lambda: [0, 0])
    u_sebestoimost = defaultdict(list)

    for sublist in data:
        for item in sublist:
            tovari_m.add(item[0])

    for item in data[0]:
        name, kolvo, price = item
        nach_o[name][0] += kolvo
        nach_o[name][1] += price
    for item in data[1]:
        name, kolvo, price = item
        prihod[name][0] += kolvo
        prihod[name][1] += price
    for item in data[2]:
        name, kolvo, price = item
        rashod[name][0] += kolvo
        rashod[name][1] += price

    for item, price in sebestoimost:
        u_sebestoimost[item].append(price)
    for item, prices in u_sebestoimost.items():
        u_sebestoimost[item] = int(sum(prices) / len(prices))

    tovari = list(tovari_m)

    nach_o = [(name, kolvo, price) for name, (kolvo, price) in nach_o.items()]
    prihod = [(name, kolvo, price) for name, (kolvo, price) in prihod.items()]
    rashod = [(name, kolvo, price) for name, (kolvo, price) in rashod.items()]

    sebestoimost = sorted(u_sebestoimost.items())

    tovari.sort()

    nach_o.sort()
    prihod.sort()
    rashod.sort()

    current_index = 0
    for i in range(len(tovari)):
        while current_index < len(nach_o) and nach_o[current_index][0] < tovari[i]:
            current_index += 1
        if current_index == len(nach_o) or nach_o[current_index][0] != tovari[i]:
            nach_o.insert(current_index, (tovari[i], 0, 0))

    current_index = 0
    for i in range(len(tovari)):
        while current_index < len(prihod) and prihod[current_index][0] < tovari[i]:
            current_index += 1
        if current_index == len(prihod) or prihod[current_index][0] != tovari[i]:
            prihod.insert(current_index, (tovari[i], 0, 0))

    current_index = 0
    for i in range(len(tovari)):
        while current_index < len(rashod) and rashod[current_index][0] < tovari[i]:
            current_index += 1
        if current_index == len(rashod) or rashod[current_index][0] != tovari[i]:
            rashod.insert(current_index, (tovari[i], 0, 0))

    nach_o.sort()
    prihod.sort()
    rashod.sort()

    sebestoimost.sort()

    print(tovari)

    print(nach_o)
    print(prihod)
    print(rashod)

    print(sebestoimost)

    sebestoimost_m = []

    for i in range(len(tovari)):
        sebestoimost_m.append(sebestoimost[i][1] * rashod[i][1])

    pribil = []

    for i in range(len(tovari)):
        pribil.append(rashod[i][2] - sebestoimost_m[i])

    konech_o = []

    for i in range(len(tovari)):
        konech_o.append(nach_o[i][2] + prihod[i][2] - sebestoimost_m[i])

    sebestoimost_m = sum(sebestoimost_m)
    pribil = sum(pribil)
    konech_o = sum(konech_o)

    print(sebestoimost_m)

    print(pribil)

    print(konech_o)

    return tovari, nach_o, prihod, rashod, sebestoimost, sebestoimost_m, konech_o, pribil

daaaaaa = mega_fun()

tovari = daaaaaa[0]
nach_o = daaaaaa[1]
prihod = daaaaaa[2]
rashod = daaaaaa[3]
sebestoimost = daaaaaa[4]
sebestoimost_m = daaaaaa[5]
konech_o = daaaaaa[6]
pribil = daaaaaa[7]

print(tovari)
print(nach_o)
print(prihod)
print(rashod)
print(sebestoimost)
print(sebestoimost_m)
print(konech_o)
print(pribil)

def create_excel_list():
    wb = Workbook()
    ws = wb.active

    ws['A1'] = 'Номенклатура'
    ws['B1'] = 'начальный остаток'
    ws['C1'] = ''
    ws['D1'] = 'приход'
    ws['E1'] = ''
    ws['F1'] = 'расход'
    ws['G1'] = ''
    ws['H1'] = 'конечный остаток'
    ws['I1'] = ''

    ws['A2'] = ''
    ws['B2'] = 'кол-во'
    ws['C2'] = 'сумма'
    ws['D2'] = 'кол-во'
    ws['E2'] = 'сумма'
    ws['F2'] = 'кол-во'
    ws['G2'] = 'сумма'
    ws['H2'] = 'кол-во'
    ws['I2'] = 'сумма'

    abc = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']

    for i in range(len(db_select.tovari)):
        for j in range(9):
            if abc[j] == 'A':
                ws[f'{abc[j]}{2+i+1}'] = db_select.tovari[i]
            if abc[j] == 'B':
                ws[f'{abc[j]}{2+i+1}'] = db_select.nach_o[i][1]
            if abc[j] == 'C':
                ws[f'{abc[j]}{2+i+1}'] = db_select.nach_o[i][2]
            if abc[j] == 'D':
                ws[f'{abc[j]}{2+i+1}'] = db_select.prihod[i][1]
            if abc[j] == 'E':
                ws[f'{abc[j]}{2+i+1}'] = db_select.prihod[i][2]
            if abc[j] == 'F':
                ws[f'{abc[j]}{2+i+1}'] = db_select.rashod[i][1]
            if abc[j] == 'G':
                ws[f'{abc[j]}{2+i+1}'] = (db_select.sebestoimost[i][1] * db_select.rashod[i][1])
            if abc[j] == 'H':
                ws[f'{abc[j]}{2+i+1}'] = f'=B{2+i+1} + D{2+i+1} - F{2+i+1}'
            if abc[j] == 'I':
                ws[f'{abc[j]}{2+i+1}'] = f'=C{2+i+1} + E{2+i+1} - G{2+i+1}'

    ws[f'A{2 + len(db_select.tovari) + 1}'] = 'итого'
    ws[f'B{2 + len(db_select.tovari) + 1}'] = ''
    ws[f'C{2 + len(db_select.tovari) + 1}'] = f'{sum(item[2] for item in db_select.nach_o)}'
    ws[f'D{2 + len(db_select.tovari) + 1}'] = ''
    ws[f'E{2 + len(db_select.tovari) + 1}'] = f'{sum(item[2] for item in db_select.prihod)}'
    ws[f'F{2 + len(db_select.tovari) + 1}'] = ''
    ws[f'G{2 + len(db_select.tovari) + 1}'] = f'{db_select.sebestoimost_m}'
    ws[f'H{2 + len(db_select.tovari) + 1}'] = ''
    ws[f'I{2 + len(db_select.tovari) + 1}'] = f'{db_select.konech_o}'

    ws.merge_cells('A1:A2')
    ws.merge_cells('B1:C1')
    ws.merge_cells('D1:E1')
    ws.merge_cells('F1:G1')
    ws.merge_cells('H1:I1')

    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['B1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['D1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['F1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['H1'].alignment = Alignment(horizontal='center', vertical='center')

    ws[f'A{2 + len(db_select.tovari) + 1}'].alignment = Alignment(horizontal='right', vertical='center')

    ws[f'C{2 + len(db_select.tovari) + 1}'].alignment = Alignment(horizontal='right', vertical='center')
    ws[f'E{2 + len(db_select.tovari) + 1}'].alignment = Alignment(horizontal='right', vertical='center')
    ws[f'G{2 + len(db_select.tovari) + 1}'].alignment = Alignment(horizontal='right', vertical='center')
    ws[f'I{2 + len(db_select.tovari) + 1}'].alignment = Alignment(horizontal='right', vertical='center')

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 10

    wb.save('report.xlsx')