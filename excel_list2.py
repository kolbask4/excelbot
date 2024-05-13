from openpyxl import Workbook
from openpyxl.styles import Alignment
import db_select

def create_excel_list2():
    wb = Workbook()
    ws = wb.active

    ws['A1'] = 'Номенклатура'
    ws['B1'] = 'Количество'
    ws['C1'] = 'Сумма продажи'
    ws['D1'] = 'Себестоимость'
    ws['E1'] = 'Прибыль'

    abc = ['A', 'B', 'C', 'D', 'E']

    for i in range(len(db_select.tovari)):
        for j in range(5):
            if abc[j] == 'A':
                ws[f'{abc[j]}{1+i+1}'] = db_select.tovari[i]
            if abc[j] == 'B':
                ws[f'{abc[j]}{1+i+1}'] = db_select.rashod[i][1]
            if abc[j] == 'C':
                ws[f'{abc[j]}{1+i+1}'] = db_select.rashod[i][2]
            if abc[j] == 'D':
                ws[f'{abc[j]}{1+i+1}'] = (db_select.sebestoimost[i][1] * db_select.rashod[i][1])
            if abc[j] == 'E':
                ws[f'{abc[j]}{1+i+1}'] = f'=C{1+i+1}-D{1+i+1}'

    ws[f'A{1 + len(db_select.tovari) + 1}'] = 'итого'
    ws[f'B{1 + len(db_select.tovari) + 1}'] = ''
    ws[f'C{1 + len(db_select.tovari) + 1}'] = f'{sum(item[2] for item in db_select.rashod)}'
    ws[f'D{1 + len(db_select.tovari) + 1}'] = f'{db_select.sebestoimost_m}'
    ws[f'E{1 + len(db_select.tovari) + 1}'] = f'{db_select.pribil}'

    ws[f'A{1 + len(db_select.tovari) + 1}'].alignment = Alignment(horizontal='right', vertical='center')

    ws[f'C{1 + len(db_select.tovari) + 1}'].alignment = Alignment(horizontal='right', vertical='center')
    ws[f'D{1 + len(db_select.tovari) + 1}'].alignment = Alignment(horizontal='right', vertical='center')
    ws[f'E{1 + len(db_select.tovari) + 1}'].alignment = Alignment(horizontal='right', vertical='center')

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20

    wb.save('report2.xlsx')